import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ========== CONFIG ==========
INPUT_FILE = "D:/Website/học tập/Thuc_Hanh_Quan_Tri_Du_An/buoi_5/Áp dụng quy trình 5 bước xử lý nhanh dữ liệu tài nguyên trong bảng Dự toán dân dụng (1).xlsx"
OUTPUT_FILE = "buoi5_ketqua.xlsx"
CHART_FILE = "buoi5_chart.png"
CANDIDATE_SHEETS = ["Chiết tính", "Đơn giá chi tiết", "Tổng hợp VT"]

# ========== HÀM HỖ TRỢ ==========
def find_valid_sheet(file):
    """Quét toàn bộ sheet để tìm sheet có chứa cột liên quan đến 'Mã'."""
    xls = pd.ExcelFile(file)
    for sheet in xls.sheet_names:
        for header in range(0, 10):  # thử 10 dòng đầu
            try:
                df = pd.read_excel(file, sheet_name=sheet, header=header, nrows=20)
                cols = [str(c).strip() for c in df.columns]
                if any("mã" in c.lower() for c in cols):
                    print(f"🔎 Sheet ứng viên: {sheet}, header={header+1}, cột={cols}")
                    return sheet, header
            except Exception:
                continue
    return None, None


def style_excel(file):
    """Tô màu và format file Excel."""
    wb = load_workbook(file)
    for ws in wb.worksheets:
        # Định dạng header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        # Định dạng dữ liệu
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
    wb.save(file)

# ========== MAIN ==========
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"❌ Không tìm thấy file {INPUT_FILE}")

sheet, header = find_valid_sheet(INPUT_FILE)
if not sheet:
    raise KeyError("❌ Không tìm thấy sheet có cột 'Mã hiệu'!")

print(f"✅ Đang đọc sheet: {sheet}, header dòng {header+1}")

df = pd.read_excel(INPUT_FILE, sheet_name=sheet, header=header)

# Chuẩn hóa tên cột
df.columns = [str(c).strip() for c in df.columns]

# Xác định cột cần thiết (chấp nhận cả "Mã số" hoặc "Mã hiệu")
col_ma = next((c for c in df.columns if "mã hiệu" in c.lower() or "mã số" in c.lower()), None)
col_kl = next((c for c in df.columns if "khối lượng" in c.lower()), None)
col_tt = next((c for c in df.columns if "thành tiền" in c.lower()), None)

if not all([col_ma, col_kl, col_tt]):
    raise KeyError("❌ Thiếu cột 'Mã hiệu' / 'Khối lượng' / 'Thành tiền'!")

# Lọc dữ liệu
df_clean = df[[col_ma, col_kl, col_tt]].dropna()
df_clean[col_ma] = df_clean[col_ma].astype(str).str.strip()

# Tổng hợp
df_sum = df_clean.groupby(col_ma).agg(
    Tong_Khoi_Luong=(col_kl, "sum"),
    Tong_Thanh_Tien=(col_tt, "sum")
).reset_index().sort_values("Tong_Khoi_Luong", ascending=False)

# Xuất Excel
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df_clean.to_excel(writer, sheet_name="Du lieu goc", index=False)
    df_sum.to_excel(writer, sheet_name="Tong hop", index=False)

style_excel(OUTPUT_FILE)
print(f"📂 Đã xuất Excel: {OUTPUT_FILE}")

# Vẽ biểu đồ
top10 = df_sum.head(10)
plt.figure(figsize=(10, 6))
plt.barh(top10[col_ma], top10["Tong_Khoi_Luong"], color="skyblue")
plt.xlabel("Khối lượng")
plt.ylabel("Mã hiệu")
plt.title("Top 10 Mã hiệu theo Khối lượng")
plt.gca().invert_yaxis()
plt.tight_layout()
plt.savefig(CHART_FILE, dpi=150)
plt.close()
print(f"📊 Đã xuất biểu đồ: {CHART_FILE}")
